"""
Scraper + generador de reporte: FootyStats (equipos por % de un stat de
goles) cruzado contra los proximos partidos de Supermatch.com.uy.

Corre con Playwright (navegador headless) porque ambos sitios cargan
contenido dinamico con JavaScript (footystats via boton "Mostrar 200
Mas", Supermatch via app Angular).

Uso:
    python footystats_supermatch_report.py \
        --stat-key over05-goals \
        --stat-label "Mas de 0,5 Goles" \
        --min-pct 80 --max-pct 100 \
        --supermatch-url "https://www.supermatch.com.uy/prematch-elastic/36/betting/sports/100001" \
        --output-dir reportes_futbol
"""

import argparse
import difflib
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

FOOTYSTATS_URL = "https://footystats.org/es/stats/over-under"

COUNTRY_NAMES = {
    "br": "Brasil", "bo": "Bolivia", "kz": "Kazajistan", "us": "Estados Unidos",
    "no": "Noruega", "fi": "Finlandia", "au": "Australia", "is": "Islandia",
    "uy": "Uruguay", "ec": "Ecuador", "pe": "Peru", "py": "Paraguay",
    "cn": "China", "se": "Suecia", "ee": "Estonia", "lv": "Letonia",
    "ar": "Argentina", "kr": "Corea del Sur", "ca": "Canada", "gb": "Reino Unido",
    "es": "Espana", "it": "Italia", "de": "Alemania", "fr": "Francia",
    "pt": "Portugal", "nl": "Holanda", "tr": "Turquia", "mx": "Mexico",
}


def normalize(name: str) -> str:
    """Normaliza un nombre de equipo para comparar: mayusculas, sin
    tildes, sin sufijos comunes (FC, SC, CF, EC, etc.)."""
    name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    name = name.upper().strip()
    name = re.sub(r"[^A-Z0-9 ]", " ", name)
    stopwords = {
        "FC", "SC", "CF", "EC", "CD", "CA", "CS", "AC", "SK", "IF", "IK",
        "FK", "BK", "UBK", "CLUB", "ATLETICO", "DEPORTIVO", "ESPORTE",
        "CLUBE", "DE", "DEL", "DA", "DO", "LA", "EL", "SPORTING",
    }
    tokens = [t for t in name.split() if t not in stopwords and len(t) > 1]
    return " ".join(tokens)


def match_score(a: str, b: str) -> float:
    na, nb = normalize(a), normalize(b)
    if not na or not nb:
        return 0.0
    if na in nb or nb in na:
        return 0.9
    return difflib.SequenceMatcher(None, na, nb).ratio()


def scrape_footystats(page, stat_key: str, min_pct: int, max_pct: int):
    page.goto(FOOTYSTATS_URL, wait_until="networkidle", timeout=60000)

    button = page.locator(f'button.ajax-get-more-stats[data-table-key="{stat_key}"]')
    section = button.locator(
        "xpath=ancestor::div[contains(@class,'section') and contains(@class,'cf')][1]"
    )

    if button.count() > 0 and button.first.is_visible():
        button.first.click()
        page.wait_for_timeout(3000)
        page.wait_for_load_state("networkidle")

    rows = section.locator("table tbody tr")
    count = rows.count()
    print(f"FootyStats: {count} filas encontradas para stat_key={stat_key}")

    results = []
    for i in range(count):
        row = rows.nth(i)
        tds = row.locator("td")
        if tds.count() < 5:
            continue
        team_cell = tds.nth(1)
        team_name = team_cell.locator("a").inner_text().strip()
        flag_class = team_cell.locator("span.flag").get_attribute("class") or ""
        m = re.search(r"flag-([a-z]{2})-\d+", flag_class)
        country_code = m.group(1) if m else ""
        country = COUNTRY_NAMES.get(country_code, country_code.upper())

        pct_text = tds.nth(3).inner_text().strip()
        pct_match = re.search(r"\d+", pct_text)
        if not pct_match:
            continue
        pct = int(pct_match.group())

        next_match = tds.nth(4).inner_text().strip()

        if min_pct <= pct <= max_pct:
            results.append(
                {"equipo": team_name, "pais": country, "pct": pct, "proximo": next_match}
            )

    # sin duplicados, ordenado de mayor a menor %
    seen = {}
    for r in results:
        seen[r["equipo"]] = r
    return sorted(seen.values(), key=lambda r: r["pct"], reverse=True)


def scrape_supermatch(page, url: str):
    page.goto(url, wait_until="networkidle", timeout=60000)
    page.wait_for_selector("div.card.h-100", timeout=30000)
    page.wait_for_timeout(2000)

    cards = page.locator("div.card.h-100")
    count = cards.count()
    print(f"Supermatch: {count} partidos encontrados")

    matches = []
    for i in range(count):
        card = cards.nth(i)
        try:
            datetime_text = card.locator(
                "div.d-flex.align-items-start.col-6.px-0.text-start"
            ).first.inner_text().strip()
        except Exception:
            datetime_text = ""
        try:
            league = card.locator(
                "div.d-flex.flex-column.col-6.px-0.text-end"
            ).first.inner_text().strip()
        except Exception:
            league = ""
        teams = card.locator("div.row.team.text-center span.float-left")
        if teams.count() < 2:
            continue
        local = teams.nth(0).inner_text().strip()
        visitante = teams.nth(1).inner_text().strip()

        matches.append(
            {
                "fecha_hora": datetime_text,
                "liga": league,
                "local": local,
                "visitante": visitante,
            }
        )
    return matches


def cross_reference(footystats_rows, supermatch_matches, threshold=0.55):
    combined = []
    for m in supermatch_matches:
        local_match = max(
            footystats_rows,
            key=lambda r: match_score(r["equipo"], m["local"]),
            default=None,
        )
        local_score = match_score(local_match["equipo"], m["local"]) if local_match else 0
        visit_match = max(
            footystats_rows,
            key=lambda r: match_score(r["equipo"], m["visitante"]),
            default=None,
        )
        visit_score = match_score(visit_match["equipo"], m["visitante"]) if visit_match else 0

        combined.append(
            {
                "fecha_hora": m["fecha_hora"],
                "liga": m["liga"],
                "local": m["local"],
                "pct_local": local_match["pct"] if local_score >= threshold else "",
                "visitante": m["visitante"],
                "pct_visitante": visit_match["pct"] if visit_score >= threshold else "",
            }
        )
    return combined


def build_xlsx(footystats_rows, combined_matches, stat_label, min_pct, max_pct, output_path: Path):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "FootyStats"
    header1 = ["Equipo", "Pais", f"% {stat_label}", "Proximo Partido"]
    ws1.append(header1)
    for r in footystats_rows:
        ws1.append([r["equipo"], r["pais"], r["pct"], r["proximo"]])
    style_header(ws1, len(header1))
    widths1 = [40, 18, 14, 16]
    for i, w in enumerate(widths1, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    ws2 = wb.create_sheet("Coincidencias Supermatch")
    header2 = [
        "Fecha/Hora", "Liga (Supermatch)", "Equipo Local",
        f"% {stat_label} Local", "Equipo Visitante", f"% {stat_label} Visitante",
    ]
    ws2.append(header2)
    for m in combined_matches:
        ws2.append(
            [m["fecha_hora"], m["liga"], m["local"], m["pct_local"], m["visitante"], m["pct_visitante"]]
        )
    style_header(ws2, len(header2))
    widths2 = [22, 26, 30, 16, 30, 18]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    ws3 = wb.create_sheet("Info")
    ws3.append(["Generado", datetime.now(timezone.utc).isoformat()])
    ws3.append(["Stat", stat_label])
    ws3.append(["Rango %", f"{min_pct}-{max_pct}"])
    ws3.append(["Equipos FootyStats encontrados", len(footystats_rows)])
    ws3.append(["Partidos Supermatch encontrados", len(combined_matches)])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Guardado: {output_path}")


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--stat-key", default="over05-goals")
    parser.add_argument("--stat-label", default="Mas de 0,5 Goles")
    parser.add_argument("--min-pct", type=int, default=80)
    parser.add_argument("--max-pct", type=int, default=100)
    parser.add_argument(
        "--supermatch-url",
        default="https://www.supermatch.com.uy/prematch-elastic/36/betting/sports/100001",
    )
    parser.add_argument("--output-dir", default="reportes_futbol")
    args = parser.parse_args()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        footystats_rows = scrape_footystats(page, args.stat_key, args.min_pct, args.max_pct)

        combined_matches = []
        try:
            supermatch_matches = scrape_supermatch(page, args.supermatch_url)
            combined_matches = cross_reference(footystats_rows, supermatch_matches)
        except Exception as e:
            print(f"AVISO: no se pudo procesar Supermatch: {e}")

        browser.close()

    today = datetime.now().strftime("%Y-%m-%d")
    output_path = Path(args.output_dir) / f"reporte_{args.stat_key}_{today}.xlsx"
    build_xlsx(footystats_rows, combined_matches, args.stat_label, args.min_pct, args.max_pct, output_path)


if __name__ == "__main__":
    main()
